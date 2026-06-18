import json
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

# Fetch feedback from production API
result = subprocess.run(
    ["curl", "-s", "https://api.unipaith.co/api/v1/feedback/admin",
     "-H", "X-Ops-Token: unipaith-ops-fbx-2026",
     "-H", "Accept: application/json"],
    capture_output=True, text=True
)

rows = json.loads(result.stdout)
rows.sort(key=lambda r: r["created_at"], reverse=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "User Feedback"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1E3A5F")
alt_fill = PatternFill("solid", fgColor="F0F4F8")
border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)

headers = ["#", "Date (UTC)", "Role", "Title", "Message", "Page / Context", "User ID"]
col_widths = [5, 22, 12, 30, 60, 30, 38]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 22

for i, row in enumerate(rows, start=1):
    excel_row = i + 1
    dt = datetime.fromisoformat(row["created_at"]).strftime("%Y-%m-%d %H:%M")
    ctx = row.get("context") or {}
    page = ctx.get("path", "") if isinstance(ctx, dict) else str(ctx)

    values = [
        i,
        dt,
        row.get("role") or "",
        row.get("title") or "",
        row.get("message") or "",
        page,
        row.get("user_id") or "",
    ]
    alignments = [center, left, center, left, left, left, left]

    fill = alt_fill if i % 2 == 0 else None

    for col_idx, (val, aln) in enumerate(zip(values, alignments), start=1):
        cell = ws.cell(row=excel_row, column=col_idx, value=val)
        cell.alignment = aln
        cell.border = border
        if fill:
            cell.fill = fill

    ws.row_dimensions[excel_row].height = 40

ws.freeze_panes = "A2"

out_path = "/home/user/UniPaith_MVP/unipaith_feedback.xlsx"
wb.save(out_path)
print(f"Saved {len(rows)} rows → {out_path}")
