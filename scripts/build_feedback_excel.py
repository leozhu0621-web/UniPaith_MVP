#!/usr/bin/env python3
"""Fetch feedback from the production API and write to an Excel workbook."""

import csv
import io
import json
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not installed — run: pip install openpyxl")

API_BASE = "https://api.unipaith.co"
OPS_TOKEN = "unipaith-ops-fbx-2026"
OUTPUT_PATH = Path(__file__).parent.parent / "feedback_export.xlsx"


def fetch_json(path: str) -> list[dict]:
    url = f"{API_BASE}/api/v1{path}"
    req = urllib.request.Request(url, headers={"X-Ops-Token": OPS_TOKEN})
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read())


def fetch_csv_rows(path: str) -> list[list[str]]:
    url = f"{API_BASE}{path}"
    req = urllib.request.Request(url, headers={"X-Ops-Token": OPS_TOKEN})
    with urllib.request.urlopen(req, timeout=15) as resp:
        text = resp.read().decode()
    reader = csv.reader(io.StringIO(text))
    return list(reader)


def style_header(ws, row: int, cols: int):
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def build_feedback_sheet(wb):
    rows = fetch_json("/feedback/admin")
    ws = wb.active
    ws.title = "User Feedback"

    headers = ["#", "Date", "Role", "Title", "Message", "User ID", "Context"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20

    for i, r in enumerate(rows, 1):
        dt = r.get("created_at", "")
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M UTC")
        except Exception:
            pass
        ctx = r.get("context")
        ctx_str = json.dumps(ctx, ensure_ascii=False) if ctx else ""
        ws.append([
            i,
            dt,
            r.get("role") or "",
            r.get("title") or "",
            r.get("message") or "",
            str(r.get("user_id") or ""),
            ctx_str,
        ])
        ws.row_dimensions[i + 1].height = 40

    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A2"
    print(f"  User Feedback: {len(rows)} rows")
    return len(rows)


def main():
    print("Fetching feedback from api.unipaith.co …")
    wb = openpyxl.Workbook()
    try:
        n = build_feedback_sheet(wb)
    except Exception as e:
        sys.exit(f"Failed to fetch feedback: {e}")

    wb.save(OUTPUT_PATH)
    print(f"\nSaved → {OUTPUT_PATH}  ({n} feedback entries)")


if __name__ == "__main__":
    main()
