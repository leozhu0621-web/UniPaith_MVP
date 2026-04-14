"""
University Enrichment Script
Reads the outreach XLSX, scrapes university websites, and populates the production DB.
Usage: python scripts/enrich_universities.py --sample NYU
       python scripts/enrich_universities.py --rank-min 30 --rank-max 300
"""
import asyncio
import json
import logging
import re
import sys
import uuid
from pathlib import Path
from urllib.parse import urljoin

import httpx
import openpyxl
from bs4 import BeautifulSoup

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "unipaith-backend" / "src"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

XLSX_PATH = "/Users/leozhu/Desktop/AI/Claude/UniPaith/US_News_30-500_University_Programs_Outreach_Database.xlsx"

# --- Data Loading ---

def load_universities(rank_min=30, rank_max=300, sample_name=None):
    """Load universities and their programs from XLSX."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    # Load university master list
    ws_univ = wb["University Master List"]
    universities = {}
    for row in ws_univ.iter_rows(min_row=2, values_only=True):
        rank, name, location, inst_type, website, email, phone, visit_link, admissions_url, total_progs, total_contacts = row
        if rank is None or name is None:
            continue
        rank = int(rank)
        if sample_name and name != sample_name:
            continue
        if not sample_name and (rank < rank_min or rank > rank_max):
            continue

        # Parse location
        city, region = None, None
        if location and "," in location:
            parts = location.split(",")
            city = parts[0].strip()
            region = parts[1].strip() if len(parts) > 1 else None

        universities[name] = {
            "rank": rank,
            "name": name,
            "city": city,
            "region": region,
            "country": "United States",
            "type": "university" if inst_type == "Private" else "university",
            "inst_type_raw": inst_type,
            "website": f"https://{website}" if website and not website.startswith("http") else website,
            "email": email,
            "phone": phone,
            "admissions_url": admissions_url,
            "programs": [],
        }

    # Load programs
    ws_prog = wb["Programs"]
    for row in ws_prog.iter_rows(min_row=2, values_only=True):
        rank, univ_name, school_college, program_name, website = row
        if univ_name in universities and program_name:
            universities[univ_name]["programs"].append({
                "name": program_name,
                "department": school_college,
                "website": website,
            })

    wb.close()
    return universities


# --- Web Scraping ---

async def scrape_university(client: httpx.AsyncClient, univ: dict) -> dict:
    """Scrape a university's website for description, logo, and photos."""
    enriched = {
        "description": None,
        "logo_url": None,
        "campus_photos": [],
        "student_body_size": None,
        "campus_setting": None,
    }

    url = univ.get("website")
    if not url:
        return enriched

    try:
        resp = await client.get(url, follow_redirects=True, timeout=15)
        if resp.status_code != 200:
            logger.warning(f"  Failed to fetch {url}: {resp.status_code}")
            return enriched

        soup = BeautifulSoup(resp.text, "html.parser")

        # Extract description from meta tags
        for meta in soup.find_all("meta"):
            if meta.get("name") in ("description", "og:description") or meta.get("property") == "og:description":
                desc = meta.get("content", "")
                if desc and len(desc) > 30:
                    enriched["description"] = desc[:500]
                    break

        # Extract logo from meta/link tags
        for link in soup.find_all("link", rel=lambda x: x and "icon" in " ".join(x)):
            href = link.get("href")
            if href:
                enriched["logo_url"] = urljoin(url, href)
                break
        if not enriched["logo_url"]:
            og_image = soup.find("meta", property="og:image")
            if og_image and og_image.get("content"):
                enriched["logo_url"] = og_image["content"]

        # Look for campus images (og:image, large images)
        og_imgs = soup.find_all("meta", property="og:image")
        for og in og_imgs:
            img_url = og.get("content")
            if img_url and img_url not in enriched["campus_photos"]:
                enriched["campus_photos"].append(img_url)
                if len(enriched["campus_photos"]) >= 3:
                    break

        logger.info(f"  Scraped {url}: desc={bool(enriched['description'])}, logo={bool(enriched['logo_url'])}, photos={len(enriched['campus_photos'])}")

    except Exception as e:
        logger.warning(f"  Scrape error for {url}: {e}")

    return enriched


def infer_degree_type(program_name: str, department: str = "") -> str:
    """Infer degree type from program name."""
    name_lower = (program_name + " " + (department or "")).lower()
    if any(kw in name_lower for kw in ["phd", "doctoral", "doctorate"]):
        return "phd"
    if any(kw in name_lower for kw in ["mba", "master", "m.s.", "m.a.", "mfa", "mph", "msw", "med", "llm"]):
        return "masters"
    if any(kw in name_lower for kw in ["certificate", "certification"]):
        return "certificate"
    if any(kw in name_lower for kw in ["diploma"]):
        return "diploma"
    return "bachelors"


# --- Database Import ---

async def import_to_database(universities: dict, dry_run=False):
    """Import enriched universities and programs into the production database."""
    import os
    os.environ.setdefault("DATABASE_URL", "postgresql+asyncpg://unipaith:unipaith@localhost:5432/unipaith")
    os.environ.setdefault("COGNITO_BYPASS", "true")
    os.environ.setdefault("AI_MOCK_MODE", "true")
    os.environ.setdefault("S3_LOCAL_MODE", "true")

    from sqlalchemy import select, text
    from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
    from sqlalchemy.orm import sessionmaker

    from unipaith.models.base import Base
    from unipaith.models.institution import Institution, Program
    from unipaith.models.user import User, UserRole

    db_url = os.environ["DATABASE_URL"]
    engine = create_async_engine(db_url, echo=False)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with async_session() as db:
        created_count = 0
        program_count = 0
        skipped = 0

        for univ_name, univ in universities.items():
            # Check if already exists
            existing = await db.execute(
                select(Institution).where(Institution.name == univ_name)
            )
            if existing.scalar_one_or_none():
                logger.info(f"  SKIP {univ_name} (already exists)")
                skipped += 1
                continue

            if dry_run:
                logger.info(f"  [DRY RUN] Would create: {univ_name} ({len(univ['programs'])} programs)")
                created_count += 1
                program_count += len(univ["programs"])
                continue

            # Create a system user for this institution
            system_user = User(
                email=f"system+{uuid.uuid4().hex[:8]}@unipaith.co",
                role=UserRole.institution_admin,
                cognito_sub=f"system-{uuid.uuid4().hex}",
            )
            db.add(system_user)
            await db.flush()

            # Create institution
            enriched = univ.get("enriched", {})
            institution = Institution(
                admin_user_id=system_user.id,
                name=univ_name,
                type=univ["type"],
                country=univ["country"],
                region=univ.get("region"),
                city=univ.get("city"),
                ranking_data={"us_news_2025": univ["rank"]},
                description_text=enriched.get("description"),
                contact_email=univ.get("email"),
                logo_url=enriched.get("logo_url"),
                website_url=univ.get("website"),
                media_gallery=enriched.get("campus_photos"),
                claimed_from_source="public_catalog",
                is_verified=False,
            )
            db.add(institution)
            await db.flush()

            # Create programs
            seen = set()
            for prog in univ["programs"]:
                key = f"{prog['name']}|{prog.get('department', '')}"
                if key in seen:
                    continue
                seen.add(key)

                program = Program(
                    institution_id=institution.id,
                    program_name=prog["name"],
                    degree_type=infer_degree_type(prog["name"], prog.get("department", "")),
                    department=prog.get("department"),
                    is_published=True,
                )
                db.add(program)
                program_count += 1

            created_count += 1
            logger.info(f"  CREATED {univ_name}: {len(univ['programs'])} programs, rank={univ['rank']}")

        if not dry_run:
            await db.commit()

        logger.info(f"\n=== Summary ===")
        logger.info(f"Created: {created_count} institutions, {program_count} programs")
        logger.info(f"Skipped: {skipped} (already existed)")

    await engine.dispose()


# --- API-based Import (for production) ---

async def import_via_api(universities: dict, api_base: str, admin_token: str, batch_size: int = 10):
    """Import universities via the production API seed endpoint."""
    async with httpx.AsyncClient(
        base_url=api_base,
        headers={"Authorization": f"Bearer {admin_token}"},
        timeout=60,
    ) as client:
        # Process in batches
        items = list(universities.values())
        for i in range(0, len(items), batch_size):
            batch = items[i:i + batch_size]
            payload = {
                "institutions": [
                    {
                        "name": u["name"],
                        "type": u["type"],
                        "country": u["country"],
                        "region": u.get("region"),
                        "city": u.get("city"),
                        "website_url": u.get("website"),
                        "contact_email": u.get("email"),
                        "description_text": u.get("enriched", {}).get("description"),
                        "logo_url": u.get("enriched", {}).get("logo_url"),
                        "media_gallery": u.get("enriched", {}).get("campus_photos"),
                        "ranking_data": {"us_news_2025": u["rank"]},
                        "programs": [
                            {
                                "program_name": p["name"],
                                "degree_type": infer_degree_type(p["name"], p.get("department", "")),
                                "department": p.get("department"),
                            }
                            for p in u["programs"]
                        ],
                    }
                    for u in batch
                ]
            }

            resp = await client.post("/api/v1/internal/seed-institutions", json=payload)
            if resp.status_code == 200:
                result = resp.json()
                logger.info(
                    f"  Batch {i // batch_size + 1}: "
                    f"created={result['created']}, "
                    f"skipped={result['skipped']}, "
                    f"programs={result['total_programs']}"
                )
            else:
                logger.error(f"  Batch {i // batch_size + 1} FAILED: {resp.status_code} {resp.text[:200]}")


# --- Main ---

async def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--sample", type=str, help="Run on a single university by name")
    parser.add_argument("--rank-min", type=int, default=30)
    parser.add_argument("--rank-max", type=int, default=300)
    parser.add_argument("--dry-run", action="store_true", help="Don't write to DB")
    parser.add_argument("--no-scrape", action="store_true", help="Skip web scraping")
    parser.add_argument("--api", type=str, help="Use API import (provide base URL e.g. https://api.unipaith.co)")
    parser.add_argument("--token", type=str, help="Admin auth token for API import")
    parser.add_argument("--batch-size", type=int, default=10, help="Batch size for API import")
    args = parser.parse_args()

    logger.info("Loading universities from XLSX...")
    universities = load_universities(args.rank_min, args.rank_max, args.sample)
    logger.info(f"Loaded {len(universities)} universities")

    total_programs = sum(len(u["programs"]) for u in universities.values())
    logger.info(f"Total programs: {total_programs}")

    if not args.no_scrape:
        logger.info("\nScraping university websites...")
        async with httpx.AsyncClient(
            headers={"User-Agent": "UniPaith-Bot/1.0 (contact: dev@unipaith.co)"},
            verify=False,
        ) as client:
            for name, univ in universities.items():
                logger.info(f"Scraping {name}...")
                enriched = await scrape_university(client, univ)
                univ["enriched"] = enriched

    if args.api:
        if not args.token:
            logger.error("--token is required for API import")
            return
        logger.info(f"\nImporting via API ({args.api})...")
        await import_via_api(universities, args.api, args.token, args.batch_size)
    else:
        logger.info("\nImporting to local database...")
        await import_to_database(universities, dry_run=args.dry_run)


if __name__ == "__main__":
    asyncio.run(main())
