#!/usr/bin/env python3
"""Export UniPaith feedback to Excel."""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

RAW = '[{"id":"d83a2738-77ad-4ca0-9584-7863375bcaaa","user_id":"044d98e6-309f-444c-b717-da3a95a19e97","role":"student","title":null,"message":"fdfd","context":{"path":"/s/explore"},"created_at":"2026-06-12T22:19:31.074734+00:00"},{"id":"bc71ba7c-6896-49da-a519-ad7d31646aae","user_id":"6c2c486d-e6d1-4aa3-a7f6-568c1848016d","role":"student","title":"Random test #3651","message":"🏒 Random test note 3651 — quokkas are the happiest animal on earth, and this feedback form finally works end-to-end. Safe to delete.","context":{"path":"/s","source":"claude-verification"},"created_at":"2026-06-06T22:07:31.549949+00:00"},{"id":"b3017060-ae51-493b-bdbb-053f0a4a8f74","user_id":"3b63296a-03cf-4c4e-9a9d-a999e30b8a7b","role":"student","title":null,"message":"sdsd","context":{"path":"/s/posts"},"created_at":"2026-06-06T22:06:36.322974+00:00"},{"id":"f5c372d5-70c5-45e7-aae9-26b817316663","user_id":"3b826e8a-3590-44df-b4d1-15a6708b7ff7","role":"student","title":null,"message":"id-token control","context":null,"created_at":"2026-06-06T20:52:43.638092+00:00"},{"id":"df77cff3-585e-415d-a1c4-dcb67430de84","user_id":"3b826e8a-3590-44df-b4d1-15a6708b7ff7","role":"student","title":null,"message":"access-token test","context":null,"created_at":"2026-06-06T20:52:43.461294+00:00"},{"id":"5d002845-2c70-4790-a7e0-16f99df0224f","user_id":"3b826e8a-3590-44df-b4d1-15a6708b7ff7","role":"student","title":null,"message":"refresh-path test","context":null,"created_at":"2026-06-06T20:44:36.701395+00:00"},{"id":"17cfa424-d896-4b47-8768-50ff54125939","user_id":"a5eaa719-0ccd-498a-ab8b-9077c93a4c6e","role":"student","title":"Claude post-deploy check","message":"✅ Post-deploy verify: invalid token now 401, valid token still 201.","context":{"path":"/s","source":"claude-verification"},"created_at":"2026-06-06T20:34:40.126481+00:00"},{"id":"eb963a76-29b4-41b2-8330-3c9ea3988b29","user_id":"a5eaa719-0ccd-498a-ab8b-9077c93a4c6e","role":"student","title":"Claude end-to-end test","message":"✅ Automated test from Claude verifying the feedback pipeline works after the Modal + 401 fixes. Safe to delete.","context":{"path":"/s","source":"claude-verification"},"created_at":"2026-06-06T20:19:53.896234+00:00"}]'

rows = json.loads(RAW)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Feedback"

# Styles
header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill("solid", fgColor="F0F4F8")
thin = Side(style="thin", color="D0D7DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top")

headers = ["#", "Date (UTC)", "Role", "Page / Context", "Title", "Message", "User ID", "Entry ID"]
col_widths = [4, 22, 10, 22, 28, 60, 38, 38]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 20
ws.freeze_panes = "A2"

for i, fb in enumerate(rows, 1):
    dt = fb["created_at"]
    ctx = fb.get("context") or {}
    page = ctx.get("path", "") if isinstance(ctx, dict) else ""

    values = [
        i,
        dt.replace("T", " ").replace("+00:00", "").split(".")[0],
        fb.get("role") or "",
        page,
        fb.get("title") or "",
        fb.get("message") or "",
        fb.get("user_id") or "",
        fb.get("id") or "",
    ]

    fill = alt_fill if i % 2 == 0 else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=i + 1, column=col, value=val)
        cell.border = border
        cell.alignment = wrap if col in (5, 6) else center
        if fill:
            cell.fill = fill

    ws.row_dimensions[i + 1].height = max(30, min(120, 15 * (len(fb.get("message") or "") // 80 + 1)))

out = "/home/user/UniPaith_MVP/unipaith_feedback.xlsx"
wb.save(out)
print(f"Saved {len(rows)} rows → {out}")
